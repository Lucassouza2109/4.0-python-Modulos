# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# openpyxl - criando uma planilha do Excel (Workbook e Worksheet)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
# PARA TRABALHAR COM CAMINHOS. 

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
# WORKSHEET = Utilizado para manipular dados. 
# Permite : Adicionar, remover, deletar


ROOT_FOLDER = Path(__file__).parent # PASTA RAIZ. 
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'
# WORKBOOK = Arquivo Excel. 

workbook = Workbook()
# worksheet: Worksheet = workbook.active # type: ignore

# ------------------------

# Nome para a planilha
sheet_name = 'Minha planilha'

# Criamos a planilha
workbook.create_sheet(sheet_name, 0)

# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]  # type: ignore

# Remover uma planilha
workbook.remove(workbook['Sheet'])

# -------------------------

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')  # Na linha | Coluna  = Nome
worksheet.cell(1, 2, 'Idade') # Na linha | Coluna  = Idade
worksheet.cell(1, 3, 'Nota')  # Na linha | Coluna  = Nota    

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    print (student)
    worksheet.append(student) 
    # VAI ADICIONAR LINHA POR LINHA NO MEU ARQUIVO.

workbook.save(WORKBOOK_PATH)
print ()
print ('Arquivo Salvo')